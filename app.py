# app.py
import streamlit as st
import pandas as pd
import sqlite3
import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Inicializa banco SQLite
DB_PATH = 'fatura.db'
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS transacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        descricao_base TEXT,
        valor REAL,
        parcelas TEXT,
        beneficiario TEXT,
        categoria TEXT
    )''')
    conn.commit()
    return conn, cursor

def extrair_parcela(desc):
    match = re.search(r'Parcela\s+(\d+)/(\d+)', desc, re.IGNORECASE)
    return f"{match.group(1)}/{match.group(2)}" if match else ''

def gerar_excel(df, caminho, lista_pessoas, total_individual):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fatura Dividida"

    header = ['Data', 'Descrição', 'Parcela', 'Valor', 'Dividido por', 'Valor por Pessoa']
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DCE6F1', fill_type='solid')

    for _, row in df.iterrows():
        ws.append([
            row['Data'].strftime('%d/%m/%Y'),
            row['Descrição'],
            row['Parcela'],
            row['Valor'],
            ", ".join(row['DivididoPor']),
            row['ValorPorPessoa']
        ])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Criar aba Totais
    ws_totais = wb.create_sheet("Totais")
    ws_totais.append(["Pessoa", "Total (R$)"])
    for pessoa, total in total_individual.items():
        ws_totais.append([pessoa, total])
    for col in range(1, 3):
        ws_totais.column_dimensions[get_column_letter(col)].width = 20

    # Criar uma aba por pessoa
    for pessoa in lista_pessoas:
        ws_pessoa = wb.create_sheet(title=pessoa[:31])  # Excel limita nome da aba a 31 caracteres
        ws_pessoa.append(['Data', 'Descrição', 'Parcela', 'Valor total', 'Participantes', 'Valor pago por essa pessoa'])

        for _, row in df.iterrows():
            if pessoa in row['DivididoPor']:
                ws_pessoa.append([
                    row['Data'].strftime('%d/%m/%Y'),
                    row['Descrição'],
                    row['Parcela'],
                    row['Valor'],
                    ", ".join(row['DivididoPor']),
                    row['ValorPorPessoa']
                ])

        for col in range(1, 7):
            ws_pessoa.column_dimensions[get_column_letter(col)].width = 20

    wb.save(caminho)

# Interface Streamlit
st.set_page_config(page_title="Divisor de Fatura", layout="wide")
st.title("🧾 Divisão de Fatura de Cartão")


# 🔹 Instruções para exportar a fatura da Nubank
with st.expander("📋 Como exportar sua fatura Nubank em CSV"):
    st.markdown("""
    ### Passo a passo:
    1. **Acesse o aplicativo Nubank:** Abra o app Nubank no seu celular.
    2. **Resumo de Faturas:** Vá em **Cartão de Crédito** > **Resumo de Faturas**.
    3. **Selecione o mês desejado:** Escolha o mês da fatura que você deseja exportar.
    4. **Ícone de envio:** Toque no ícone de envio (no canto superior direito da tela).
    5. **Escolha o formato CSV:** Selecione **CSV** como formato de exportação.

    ### ℹ️ Observações:
    - A opção de exportar em CSV **só aparece para faturas já fechadas**.
    - O arquivo CSV será enviado por e-mail para o endereço **cadastrado na sua conta Nubank**.
    """)

# Upload do arquivo CSV
file = st.file_uploader("📤 Faça upload da fatura (.csv com colunas 'date', 'title', 'amount')", type="csv")

if file:
    df = pd.read_csv(file)
    df.rename(columns={'date': 'Data', 'title': 'Descrição', 'amount': 'Valor'}, inplace=True)
    df['Data'] = pd.to_datetime(df['Data'])
    df['Parcela'] = df['Descrição'].apply(extrair_parcela)

    conn, cursor = init_db()

    st.subheader("📌 Classifique e divida os itens")
    pessoas = st.text_input("👥 Informe os nomes das pessoas para divisão, separados por vírgula:")
    lista_pessoas = [p.strip() for p in pessoas.split(',') if p.strip()]

    dividido_por = []
    valor_por_pessoa = []

    for i, row in df.iterrows():
        descricao = row['Descrição']
        valor = row['Valor']
        parcela = row['Parcela']
        base = f"{descricao}|{valor}|{parcela}"

        cursor.execute("SELECT beneficiario, categoria FROM transacoes WHERE descricao_base=?", (base,))
        resultado = cursor.fetchone()

        with st.expander(f"{descricao} | R$ {valor:.2f} | Parcela: {parcela or 'N/A'}"):
            if resultado:
                st.info("Transação reconhecida de outra fatura.")
                beneficiar = resultado[0]
                categoria = resultado[1]
            else:
                beneficiar = st.selectbox("Beneficiário", lista_pessoas, key=f"b_{i}")
                categoria = st.text_input("Categoria", key=f"cat_{i}")

                if beneficiar and categoria:
                    cursor.execute("INSERT INTO transacoes (descricao_base, valor, parcelas, beneficiario, categoria) VALUES (?, ?, ?, ?, ?)",
                        (base, valor, parcela, beneficiar, categoria))
                    conn.commit()

            dividir_com = st.multiselect("Dividir com", options=lista_pessoas, default=[beneficiar], key=f"dividir_{i}")
            dividido_por.append(dividir_com)

            valor_unitario = round(valor / len(dividir_com), 2) if dividir_com else 0.0
            valor_por_pessoa.append(valor_unitario)
            st.markdown(f"**💰 Cada um paga: R$ {valor_unitario:.2f}**")

    df['DivididoPor'] = dividido_por
    df['ValorPorPessoa'] = valor_por_pessoa

    st.subheader("📊 Total por pessoa")
    total_individual = {}
    for index, row in df.iterrows():
        for pessoa in row['DivididoPor']:
            total_individual[pessoa] = total_individual.get(pessoa, 0) + row['ValorPorPessoa']

    total_df = pd.DataFrame.from_dict(total_individual, orient='index', columns=['Total (R$)'])
    st.dataframe(total_df.style.format({'Total (R$)': 'R$ {:.2f}'}))

    # 🧾 Detalhamento individual
    st.subheader("🧾 Detalhamento do que cada pessoa está pagando")
    for pessoa in lista_pessoas:
        st.markdown(f"### 👤 {pessoa}")
        dados = []
        for _, row in df.iterrows():
            if pessoa in row['DivididoPor']:
                dados.append({
                    'Data': row['Data'].strftime('%d/%m/%Y'),
                    'Descrição': row['Descrição'],
                    'Parcela': row['Parcela'],
                    'Valor total': f"R$ {row['Valor']:.2f}",
                    'Participantes': ", ".join(row['DivididoPor']),
                    'Valor pago por essa pessoa': f"R$ {row['ValorPorPessoa']:.2f}"
                })
        if dados:
            detalhado_df = pd.DataFrame(dados)
            st.dataframe(detalhado_df)
        else:
            st.write("Nenhuma transação atribuída a esta pessoa.")

    st.subheader("📥 Gerar planilha")
    nome_excel = f"fatura_dividida_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if st.button("📁 Gerar Excel"):
        gerar_excel(df, nome_excel, lista_pessoas, total_individual)
        with open(nome_excel, 'rb') as f:
            st.download_button("⬇️ Baixar Excel", f, file_name=nome_excel)
        os.remove(nome_excel)

    conn.close()
